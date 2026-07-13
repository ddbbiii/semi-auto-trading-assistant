from __future__ import annotations

from csv import DictReader
from datetime import datetime, timezone
from io import BytesIO, StringIO
import re
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook
from PIL import Image, ImageOps
import pytesseract
from pytesseract import Output

from .schemas import HoldingInput, ImportPreview


ALIASES = {
    "symbol": ("symbol", "代码", "证券代码", "股票代码"),
    "name": ("name", "名称", "证券名称", "股票名称"),
    "market": ("market", "市场"),
    "quantity": ("quantity", "数量", "持仓数量", "总持仓"),
    "available_quantity": ("available_quantity", "可用数量", "可卖数量"),
    "currency": ("currency", "币种"),
    "market_value": ("market_value", "市值", "持仓市值"),
    "price": ("price", "现价", "最新价"),
    "average_cost": ("average_cost", "成本", "成本价", "平均成本"),
    "holding_pnl": ("holding_pnl", "持仓盈亏", "浮动盈亏"),
    "holding_pnl_percent": ("holding_pnl_percent", "盈亏比例", "盈亏%"),
}


def preview_import(file_name: str, content_type: str | None, data: bytes) -> ImportPreview:
    if len(data) > 10 * 1024 * 1024:
        raise ValueError("文件不能超过 10 MB。")
    suffix = file_name.lower().rsplit(".", 1)[-1] if "." in file_name else ""
    warnings: list[str] = []
    account: dict[str, Any] = {}
    parser = suffix or content_type or "unknown"
    rows: list[dict[str, Any]]
    if suffix == "csv":
        text = data.decode("utf-8-sig")
        rows = list(DictReader(StringIO(text)))
    elif suffix in {"xlsx", "xlsm"}:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        headers = [str(value or "").strip() for value in values[0]] if values else []
        rows = [dict(zip(headers, row, strict=False)) for row in values[1:]]
    elif suffix in {"png", "jpg", "jpeg", "webp"} or (content_type or "").startswith("image/"):
        image = Image.open(BytesIO(data)).convert("RGB")
        ocr_data = pytesseract.image_to_data(
            image,
            lang="chi_sim+eng",
            config="--psm 11",
            output_type=Output.DICT,
        )
        rows, layout_warnings = _rows_from_broker_layout(image, ocr_data)
        declared_count = _declared_holding_count(_ocr_tokens(ocr_data))
        if declared_count:
            account["declared_holding_count"] = declared_count
        warnings.extend(layout_warnings)
        if not rows:
            text = pytesseract.image_to_string(image, lang="chi_sim+eng", config="--psm 6")
            rows = _rows_from_ocr(text)
        warnings.append("截图 OCR 结果必须逐项确认；未识别字段请手工补齐。")
    else:
        raise ValueError("仅支持 PNG/JPG/WebP、CSV 和 XLSX。")

    holdings: list[HoldingInput] = []
    for index, row in enumerate(rows, start=1):
        try:
            holdings.append(_holding_from_row(row))
        except (ValueError, TypeError) as exc:
            warnings.append(f"第 {index} 行未导入：{exc}")
    if not holdings:
        warnings.append("没有自动识别出完整持仓，请在确认页手工添加。")
    declared_count = account.get("declared_holding_count")
    if isinstance(declared_count, int) and len(holdings) < declared_count:
        warnings.append(
            f"截图显示账户共 {declared_count} 条持仓，本次只识别出 {len(holdings)} 条完整记录；"
            f"还缺 {declared_count - len(holdings)} 条，请补充未完整显示的截图。"
        )
    return ImportPreview(
        import_id=uuid4().hex,
        file_name=file_name,
        parser=parser,
        account=account,
        holdings=holdings,
        warnings=warnings,
    )


def preview_import_batch(files: list[tuple[str, str | None, bytes]]) -> ImportPreview:
    if not files:
        raise ValueError("至少需要一个文件。")
    if len(files) > 8:
        raise ValueError("一次最多识别 8 张截图。")
    if len(files) > 1 and any(not _is_image_file(name, content_type) for name, content_type, _ in files):
        raise ValueError("多文件导入只支持截图；CSV 或 XLSX 请单独上传。")

    previews = [preview_import(name, content_type, data) for name, content_type, data in files]
    if len(previews) == 1:
        return previews[0]

    holdings_by_identity: dict[tuple[str, str], HoldingInput] = {}
    warnings: list[str] = []
    declared_counts = [
        count
        for preview in previews
        if isinstance((count := preview.account.get("declared_holding_count")), int)
    ]
    for preview in previews:
        warnings.extend(
            f"{preview.file_name}：{warning}"
            for warning in preview.warnings
            if not warning.startswith("截图显示账户共")
        )
        for holding in preview.holdings:
            identity = (holding.market, holding.symbol.upper())
            existing = holdings_by_identity.get(identity)
            if existing is None:
                holdings_by_identity[identity] = holding
                continue
            if existing.model_dump() != holding.model_dump():
                warnings.append(f"{holding.symbol} 在多张截图中的数据不一致，确认表暂保留首次识别结果，请手工核对。")
            else:
                warnings.append(f"{holding.symbol} 在多张截图中重复出现，已自动合并。")

    if not holdings_by_identity:
        warnings.append("所有截图均未自动识别出完整持仓，请在确认页手工添加。")
    declared_count = max(declared_counts, default=0)
    if declared_count and len(holdings_by_identity) < declared_count:
        warnings.append(
            f"截图显示账户共 {declared_count} 条持仓，合并后只识别出 {len(holdings_by_identity)} 条完整记录；"
            f"还缺 {declared_count - len(holdings_by_identity)} 条，请补充未完整显示的截图。"
        )
    return ImportPreview(
        import_id=uuid4().hex,
        file_name=f"{len(previews)} 张账户截图",
        parser="multi_image_ocr",
        account={"declared_holding_count": declared_count} if declared_count else {},
        holdings=list(holdings_by_identity.values()),
        warnings=list(dict.fromkeys(warnings)),
    )


def snapshot_payload(request: Any) -> dict[str, Any]:
    return {
        "as_of": request.as_of.astimezone(timezone.utc).isoformat(),
        "source": request.source,
        "account": request.account,
        "pending_order_count": request.pending_order_count,
        "pending_orders": [],
        "holdings": [
            {
                **holding.model_dump(),
                "screenshot_price": holding.price,
                "monitor_priority": "normal",
            }
            for holding in request.holdings
        ],
    }


def _holding_from_row(row: dict[str, Any]) -> HoldingInput:
    normalized = {str(key).strip(): value for key, value in row.items()}
    symbol = str(_pick(normalized, "symbol") or "").strip().upper()
    if not symbol:
        raise ValueError("缺少证券代码")
    market = str(_pick(normalized, "market") or _market_for(symbol)).upper()
    currency = str(_pick(normalized, "currency") or _currency_for(market)).upper()
    return HoldingInput(
        symbol=symbol,
        name=str(_pick(normalized, "name") or ""),
        market=market,  # type: ignore[arg-type]
        quantity=_number(_pick(normalized, "quantity"), "数量"),
        available_quantity=_optional_number(_pick(normalized, "available_quantity")),
        currency=currency,  # type: ignore[arg-type]
        market_value=_number(_pick(normalized, "market_value"), "市值"),
        price=_number(_pick(normalized, "price"), "现价"),
        average_cost=_number(_pick(normalized, "average_cost"), "成本价"),
        holding_pnl=_optional_number(_pick(normalized, "holding_pnl")),
        holding_pnl_percent=_optional_number(_pick(normalized, "holding_pnl_percent")),
    )


def _rows_from_ocr(text: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    pattern = re.compile(
        r"(?P<symbol>(?:\d{5}\.HK|\d{6}\.(?:SZ|SH|SS)|[A-Z]{1,6}))\s+"
        r"(?P<name>\S+)\s+(?P<quantity>[\d,.]+)\s+(?P<market_value>[\d,.]+)\s+"
        r"(?P<price>[\d,.]+)\s+(?P<average_cost>[\d,.]+)",
        re.IGNORECASE,
    )
    for line in text.splitlines():
        match = pattern.search(line.replace("，", ","))
        if match:
            rows.append(match.groupdict())
    return rows


def _rows_from_broker_layout(image: Image.Image, data: dict[str, list[Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    tokens = _ocr_tokens(data)
    width, height = image.size
    headers = [token for token in tokens if token["left"] < width * 0.3 and "代码" in token["text"]]
    if not headers:
        return [], []

    rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    seen_symbols: set[str] = set()
    for header in headers:
        header_y = header["top"]
        market = _market_before_header(tokens, header_y)
        next_section = min(
            (
                token["top"]
                for token in tokens
                if token["top"] > header_y + height * 0.08 and _market_from_label(token["text"])
            ),
            default=height * 0.94,
        )
        candidates = [
            token
            for token in tokens
            if header_y + height * 0.02 < token["top"] < next_section
            and token["left"] < width * 0.23
            and re.fullmatch(r"[A-Za-z]{2,6}|\d{4,6}", re.sub(r"[^A-Za-z0-9]", "", token["text"]))
        ]
        for candidate in candidates:
            code_y = candidate["top"]
            row_gap = height * 0.04
            top_band = (code_y - row_gap, code_y - height * 0.007)
            code_band = (code_y - height * 0.01, code_y + height * 0.022)
            name_box = (width * 0.03, top_band[0], width * 0.32, top_band[1])
            name = _best_name(
                _text_from_tokens(tokens, name_box),
                _ocr_crop(image, name_box, lang="chi_sim+eng"),
            )
            if not name:
                continue
            raw_code = _ocr_crop(
                image,
                (width * 0.03, code_band[0], width * 0.23, code_band[1]),
                lang="eng",
                whitelist="ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789.",
            )
            symbol = _normalize_symbol(raw_code or candidate["text"], market)
            if not symbol or symbol in seen_symbols:
                continue

            quantity_values = _field_values(
                image,
                tokens,
                (width * 0.35, code_band[0], width * 0.60, code_band[1]),
                decimals=0,
            )
            market_value_values = _field_values(
                image,
                tokens,
                (width * 0.31, top_band[0], width * 0.60, top_band[1]),
                decimals=2,
            )
            price_values = _field_values(
                image,
                tokens,
                (width * 0.60, top_band[0], width * 0.84, top_band[1]),
                decimals=3,
            )
            cost_values = _field_values(
                image,
                tokens,
                (width * 0.60, code_band[0], width * 0.84, code_band[1]),
                decimals=3,
            )
            if not all((quantity_values, market_value_values, price_values, cost_values)):
                continue

            quantity = quantity_values[0]
            market_value, price, reconciliation_warning = _reconcile_position_values(
                symbol,
                quantity,
                market_value_values,
                price_values,
            )
            if reconciliation_warning:
                warnings.append(reconciliation_warning)
            rows.append(
                {
                    "symbol": symbol,
                    "name": name,
                    "market": market or _market_for(symbol),
                    "currency": _currency_for(market or _market_for(symbol)),
                    "quantity": quantity,
                    "market_value": market_value,
                    "price": price,
                    "average_cost": cost_values[0],
                }
            )
            seen_symbols.add(symbol)
    return rows, warnings


def _ocr_tokens(data: dict[str, list[Any]]) -> list[dict[str, Any]]:
    tokens: list[dict[str, Any]] = []
    for index, value in enumerate(data.get("text", [])):
        text = str(value).strip()
        if not text:
            continue
        tokens.append(
            {
                "text": text,
                "left": int(data["left"][index]),
                "top": int(data["top"][index]),
                "width": int(data["width"][index]),
                "height": int(data["height"][index]),
                "confidence": float(data["conf"][index]),
            }
        )
    return tokens


def _field_values(
    image: Image.Image,
    tokens: list[dict[str, Any]],
    box: tuple[float, float, float, float],
    *,
    decimals: int,
) -> list[float]:
    left, top, right, bottom = box
    values: list[float] = []
    matching = sorted(
        (
            token
            for token in tokens
            if left <= token["left"] + token["width"] / 2 <= right
            and top <= token["top"] + token["height"] / 2 <= bottom
        ),
        key=lambda token: (-token["confidence"], token["left"]),
    )
    for token in matching:
        parsed = _parse_ocr_number(token["text"], decimals=decimals)
        if parsed is not None:
            values.append(parsed)
    cropped = _ocr_crop(image, box, lang="eng", whitelist="0123456789.,-")
    parsed_crop = _parse_ocr_number(cropped, decimals=decimals)
    if parsed_crop is not None:
        values.append(parsed_crop)
    return values


def _ocr_crop(
    image: Image.Image,
    box: tuple[float, float, float, float],
    *,
    lang: str,
    whitelist: str | None = None,
) -> str:
    width, height = image.size
    normalized = (
        max(0, int(box[0])),
        max(0, int(box[1])),
        min(width, int(box[2])),
        min(height, int(box[3])),
    )
    crop = ImageOps.autocontrast(ImageOps.grayscale(image.crop(normalized)))
    config = "--psm 7"
    if whitelist:
        config += f" -c tessedit_char_whitelist={whitelist}"
    return pytesseract.image_to_string(crop, lang=lang, config=config).strip()


def _parse_ocr_number(value: str, *, decimals: int) -> float | None:
    cleaned = value.replace("O", "0").replace("o", "0").replace("，", ",").strip()
    match = re.search(r"-?\d[\d,]*(?:\.\d+)?", cleaned)
    if not match:
        return None
    token = match.group(0).replace(",", "")
    if decimals and "." not in token:
        digits = token.removeprefix("-")
        if len(digits) > decimals + 1:
            sign = -1 if token.startswith("-") else 1
            return sign * int(digits) / (10**decimals)
    return float(token)


def _reconcile_position_values(
    symbol: str,
    quantity: float,
    market_values: list[float],
    prices: list[float],
) -> tuple[float, float, str | None]:
    if quantity <= 0:
        return market_values[0], prices[0], None
    best = min(
        ((market_value, price) for market_value in market_values for price in prices),
        key=lambda pair: abs(pair[0] - quantity * pair[1]) / max(abs(pair[0]), abs(quantity * pair[1]), 1),
    )
    error = abs(best[0] - quantity * best[1]) / max(abs(best[0]), abs(quantity * best[1]), 1)
    if error <= 0.02:
        return best[0], best[1], None

    price_spread = max(prices) - min(prices)
    if price_spread / max(abs(prices[0]), 1e-9) <= 0.01:
        price = prices[0]
        corrected_market_value = round(quantity * price, 4)
        return (
            corrected_market_value,
            price,
            f"{symbol} 的市值 OCR 与数量×现价不一致，确认表暂按数量×现价计算，请核对原图。",
        )
    implied_price = round(market_values[0] / quantity, 4)
    return (
        market_values[0],
        implied_price,
        f"{symbol} 的现价 OCR 与市值/数量不一致，确认表暂按市值÷数量计算，请核对原图。",
    )


def _market_before_header(tokens: list[dict[str, Any]], header_y: float) -> str:
    labels = [token for token in tokens if token["top"] < header_y and _market_from_label(token["text"])]
    if not labels:
        return ""
    return _market_from_label(max(labels, key=lambda token: token["top"])["text"])


def _declared_holding_count(tokens: list[dict[str, Any]]) -> int | None:
    for token in tokens:
        if "持仓" not in token["text"]:
            continue
        nearby = sorted(
            (
                candidate
                for candidate in tokens
                if candidate["left"] >= token["left"]
                and abs(candidate["top"] - token["top"]) <= 30
                and candidate["left"] - token["left"] <= 260
            ),
            key=lambda candidate: candidate["left"],
        )
        joined = "".join(candidate["text"] for candidate in nearby)
        if match := re.search(r"[（(]?([1-9]\d{0,2})[）)]", joined):
            return int(match.group(1))
    return None


def _text_from_tokens(tokens: list[dict[str, Any]], box: tuple[float, float, float, float]) -> str:
    left, top, right, bottom = box
    matching = sorted(
        (
            token
            for token in tokens
            if left <= token["left"] + token["width"] / 2 <= right
            and top <= token["top"] + token["height"] / 2 <= bottom
        ),
        key=lambda token: token["left"],
    )
    return "".join(token["text"] for token in matching)


def _market_from_label(value: str) -> str:
    if "港股" in value:
        return "HK"
    if "美股" in value:
        return "US"
    if "A股" in value or "沪深" in value:
        return "CN"
    return ""


def _normalize_symbol(value: str, market: str) -> str:
    if market == "HK" and (match := re.search(r"\d{5}", value)):
        return f"{match.group(0)}.HK"
    if market == "CN" and (match := re.search(r"\d{6}", value)):
        return f"{match.group(0)}.SZ"
    if market == "US" and (match := re.search(r"[A-Za-z]{1,6}", value)):
        return match.group(0).upper()
    cleaned = re.sub(r"[^A-Za-z0-9.]", "", value).upper()
    if market == "HK" and re.fullmatch(r"\d{5}", cleaned):
        return f"{cleaned}.HK"
    if market == "CN" and re.fullmatch(r"\d{6}", cleaned):
        return f"{cleaned}.SZ"
    if market == "US" and re.fullmatch(r"[A-Z]{1,6}", cleaned):
        return cleaned
    if re.fullmatch(r"\d{5}", cleaned):
        return f"{cleaned}.HK"
    if re.fullmatch(r"[A-Z]{1,6}", cleaned):
        return cleaned
    return ""


def _clean_name(value: str) -> str:
    cleaned = re.sub(r"\s+", "", value).strip(".…,，|()（）")
    return cleaned if len(cleaned) >= 2 else ""


def _best_name(*values: str) -> str:
    cleaned = [_clean_name(value) for value in values]
    return max(cleaned, key=lambda value: (len(re.findall(r"[\u4e00-\u9fff]", value)), len(value)), default="")


def _is_image_file(file_name: str, content_type: str | None) -> bool:
    suffix = file_name.lower().rsplit(".", 1)[-1] if "." in file_name else ""
    return suffix in {"png", "jpg", "jpeg", "webp"} or (content_type or "").startswith("image/")


def _pick(row: dict[str, Any], field: str) -> Any:
    for alias in ALIASES[field]:
        if alias in row and row[alias] not in (None, ""):
            return row[alias]
    return None


def _number(value: Any, label: str) -> float:
    parsed = _optional_number(value)
    if parsed is None:
        raise ValueError(f"缺少{label}")
    return parsed


def _optional_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    cleaned = str(value).replace(",", "").replace("%", "").strip()
    return float(cleaned)


def _market_for(symbol: str) -> str:
    if symbol.endswith(".HK"):
        return "HK"
    if symbol.endswith((".SZ", ".SH", ".SS")) or symbol.isdigit():
        return "CN"
    return "US"


def _currency_for(market: str) -> str:
    return {"HK": "HKD", "CN": "CNY"}.get(market, "USD")
